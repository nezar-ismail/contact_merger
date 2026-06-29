import csv
from pathlib import Path

from openpyxl import load_workbook

from models import Contact

from utils import (
    contact_score,
    detect_phone_columns,
    detect_email_columns,
)

from normalizer import (
    phone_fingerprints,
)


def load_contacts(filename):
    filename = Path(filename)

    if filename.suffix.lower() == ".csv":
        return load_csv_contacts(filename)

    return load_xlsx_contacts(filename)


def read_headers(filename):
    filename = Path(filename)

    if filename.suffix.lower() == ".csv":
        with filename.open(
            newline="",
            encoding="utf-8-sig",
        ) as handle:
            reader = csv.reader(handle)
            return next(reader)

    workbook = load_workbook(
        filename,
        read_only=True,
    )
    worksheet = workbook.active

    return [
        cell.value
        for cell in worksheet[1]
    ]


def load_xlsx_contacts(filename):
    wb = load_workbook(filename)

    ws = wb.active

    headers = [
        c.value
        for c in ws[1]
    ]

    phone_cols = detect_phone_columns(headers)

    email_cols = detect_email_columns(headers)

    contacts = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2,
    ):

        fields = {}

        for h, value in zip(headers, row):
            fields[h] = value

        contact = Contact(

            row=row_index,

            score=contact_score(row),

            fields=fields,

        )

        ###################################################
        # Phones
        ###################################################

        for col in phone_cols:

            phone = row[col]

            if phone is None:
                continue

            phone = str(phone).strip()

            if phone == "":
                continue

            label = headers[col].replace(
                "Value",
                "Label",
            )

            label = fields.get(label, "")

            contact.phones.append(
                (
                    label,
                    phone,
                )
            )

            contact.phone_fingerprints.update(
                phone_fingerprints(phone)
            )

        ###################################################
        # Emails
        ###################################################

        for col in email_cols:

            email = row[col]

            if email is None:
                continue

            email = str(email).strip()

            if email == "":
                continue

            label = headers[col].replace(
                "Value",
                "Label",
            )

            label = fields.get(label, "")

            contact.emails.append(
                (
                    label,
                    email,
                )
            )

        if fields.get("Organization Name"):

            contact.organizations.add(
                fields["Organization Name"]
            )

        if fields.get("Notes"):

            contact.notes.append(
                fields["Notes"]
            )

        contacts.append(contact)

    return contacts


def load_csv_contacts(filename):
    with filename.open(
        newline="",
        encoding="utf-8-sig",
    ) as handle:
        reader = csv.reader(handle)
        headers = next(reader)
        phone_cols = detect_phone_columns(headers)
        email_cols = detect_email_columns(headers)
        contacts = []

        for row_index, row in enumerate(reader, start=2):
            row = pad_row(row, len(headers))
            contacts.append(
                build_contact(
                    row_index,
                    headers,
                    row,
                    phone_cols,
                    email_cols,
                )
            )

    return contacts


def build_contact(row_index, headers, row, phone_cols, email_cols):
    fields = {}

    for h, value in zip(headers, row):
        fields[h] = value

    contact = Contact(
        row=row_index,
        score=contact_score(row),
        fields=fields,
    )

    for col in phone_cols:
        phone = row[col]

        if phone is None:
            continue

        phone = str(phone).strip()

        if phone == "":
            continue

        label_header = headers[col].replace(
            "Value",
            "Label",
        )
        label = fields.get(label_header, "")
        contact.phones.append((label, phone))
        contact.phone_fingerprints.update(phone_fingerprints(phone))

    for col in email_cols:
        email = row[col]

        if email is None:
            continue

        email = str(email).strip()

        if email == "":
            continue

        label_header = headers[col].replace(
            "Value",
            "Label",
        )
        label = fields.get(label_header, "")
        contact.emails.append((label, email))

    if fields.get("Organization Name"):
        contact.organizations.add(fields["Organization Name"])

    if fields.get("Notes"):
        contact.notes.append(fields["Notes"])

    return contact


def pad_row(row, width):
    if len(row) >= width:
        return row

    return row + [""] * (width - len(row))
